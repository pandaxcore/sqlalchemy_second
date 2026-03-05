from datetime import datetime
from openpyxl import load_workbook
from engine import engine
from model import Base, People
from sqlalchemy.orm import Session


Base.metadata.create_all(engine)

wb = load_workbook('file_example_XLS_10.xlsx')

sheet = wb['Sheet1']

with Session(engine) as session:
    for row in sheet.iter_rows(min_row=2, values_only=True):

        date_pattern = "%d/%m/%Y"
        date_obj = datetime.strptime(row[6], date_pattern)

        person = People(
            first_name=row[1],
            last_name=row[2],
            gender=row[3],
            country=row[4],
            age=row[5],
            date=date_obj
        )

        session.add(person)

    session.commit()
